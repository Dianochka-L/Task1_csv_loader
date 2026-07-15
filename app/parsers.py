"""Чтение табличных файлов
Здесь используется только чтение файла (csv-модуль и openpyxl для xlsx)
"""
from __future__ import annotations
import csv
import io
from typing import List, Tuple
from openpyxl import load_workbook
# Разделители-кандидаты в порядке приоритета при равенстве.
_CANDIDATE_DELIMITERS = (";", ",", "\t", "|")


class ParseError(ValueError):
    """Ошибка разбора входного файла (неверный формат, пустая таблица итп)"""

Table = Tuple[List[str], List[List[str]]]

def _sniff_delimiter(sample: str) -> str:
    """Определить разделитель по первой непустой строке файла
    Для каждого кандидата считаем, на сколько полей он делит строку заголовка,
    и выбираем разделитель с наибольшим числом полей. При равенстве - по порядку
    приоритета в _CANDIDATE_DELIMITERS
    """
    first_line = next((ln for ln in sample.splitlines() if ln.strip()), "")
    best, best_fields = ",", 1
    for delim in _CANDIDATE_DELIMITERS:
        fields = len(first_line.split(delim))
        if fields > best_fields:
            best, best_fields = delim, fields
    return best

def parse_csv(content: bytes) -> Table:   
    text = content.decode("utf-8-sig")
    delimiter = _sniff_delimiter(text[:4096])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        raise ParseError("Файл пустой или не содержит данных")
    headers, data = rows[0], rows[1:]
    return _normalise(headers, data)


def parse_xlsx(content: bytes) -> Table:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    raw_rows: List[List[str]] = []
    for row in rows_iter:
        cells = ["" if v is None else str(v) for v in row]
        if any(c.strip() for c in cells):
            raw_rows.append(cells)
    workbook.close()
    if not raw_rows:
        raise ParseError("Лист Excel пустой или не содержит данных")
    headers, data = raw_rows[0], raw_rows[1:]
    return _normalise(headers, data)

def _normalise(headers: List[str], data: List[List[str]]) -> Table:
    """Проверить корректность таблицы и выровнять длины строк по заголовку"""
    headers = [h.strip() for h in headers]
    if not headers or any(h == "" for h in headers):
        raise ParseError("Первая строка должна содержать непустые названия столбцов")
    if len(set(headers)) != len(headers):
        raise ParseError("Названия столбцов должны быть уникальными")
    width = len(headers)
    normalised: List[List[str]] = []
    for i, row in enumerate(data, start=2):
        if len(row) > width:
            raise ParseError(f"Строка {i}: столбцов больше, чем в заголовке")
        # Недостающие ячейки дополняем пустыми значениями (трактуются как NULL).
        row = [str(c).strip() for c in row] + [""] * (width - len(row))
        normalised.append(row)
    return headers, normalised

def parse_file(filename: str, content: bytes) -> Table:
    name = filename.lower()
    if name.endswith(".csv"):
        return parse_csv(content)
    if name.endswith(".xlsx"):
        return parse_xlsx(content)
    raise ParseError("Поддерживаются только файлы .csv и .xlsx.")